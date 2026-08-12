#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
组织健康诊断 - 数据分析脚本
读取员工问卷明细表(CSV/XLSX)，计算 公司 / 一级事业部 / 二级 / 三级部门 四级聚合指标，输出 analysis.json

用法:
  python analyze_survey.py --input 明细表.xlsx --output analysis.json
                           [--history 上期analysis.json] [--period "2026 H1"] [--min-n 4]

输入表列（按表头文字自动识别，无需固定列名）:
  组织列: 表头含「一级」「三级」(及可选「二级」) 的列自动识别为 一级事业部/二级/三级部门
  题项列: Q1..Q12 按题面关键词自动识别；缺失题自动跳过
  分数: 固定 5 分制（1-5），不做刻度换算；其余列（开放文本题、员工ID 等）自动忽略
"""
import argparse
import json
import re
import statistics
import sys
from pathlib import Path

QUESTIONS = {
    "Q1": "我知道对我的工作要求。",
    "Q2": "我有做好我的工作所需要的材料和设备。",
    "Q3": "在工作中，我每天都有机会做我最擅长做的事。",
    "Q4": "在过去七天里，我因工作出色而受到表扬。",
    "Q5": "我觉得主管或同事关心我的个人情况。",
    "Q6": "工作单位有人鼓励我的发展。",
    "Q7": "在工作中，我觉得我的意见受到重视。",
    "Q8": "公司的使命/目标使我觉得我的工作重要。",
    "Q9": "我的同事们致力于高质量的工作。",
    "Q10": "我在工作单位有一个最要好的朋友。",
    "Q11": "过去六个月内，有人和我谈及我的进步。",
    "Q12": "过去一年里，我在工作中有机会学习和成长。",
}
Q_SHORT = {
    "Q1": "知道工作要求", "Q2": "材料与设备", "Q3": "做擅长的事", "Q4": "认可与表扬",
    "Q5": "被关心", "Q6": "发展鼓励", "Q7": "意见被重视", "Q8": "感到工作重要",
    "Q9": "同事高质量", "Q10": "好友与倾诉", "Q11": "谈及进步", "Q12": "成长发展",
}
# 四层结构（盖洛普 Q12 官方四层 + 核心问题）
DIMENSIONS = {
    "基本需求": {"questions": ["Q1", "Q2"], "core_question": "我得到了什么？"},
    "管理支持": {"questions": ["Q3", "Q4", "Q5", "Q6"], "core_question": "我做了什么贡献？"},
    "团队归属": {"questions": ["Q7", "Q8", "Q9", "Q10"], "core_question": "我有归属感吗？"},
    "成长发展": {"questions": ["Q11", "Q12"], "core_question": "我如何成长？"},
}
Q_KEYWORDS = {
    "Q1": ["工作要求"],
    "Q2": ["材料", "设备"],
    "Q3": ["最擅长"],
    "Q4": ["表扬"],
    "Q5": ["关心", "个人情况"],
    "Q6": ["鼓励我的发展", "鼓励我的职业发展", "职业发展"],
    "Q7": ["意见受到重视", "我的意见受到重视"],
    "Q8": ["使命", "我的工作重要"],
    "Q9": ["高质量"],
    "Q10": ["最要好的朋友"],
    "Q11": ["谈及我的进步", "我的进步"],
    "Q12": ["学习和成长", "成长发展"],
}
# 运行时按实际数据识别题项与组织列（见 detect_schema）
QPRESENT = []    # 实际存在的题项 canonical id 列表
QDETECT = {}     # canonical id -> 原始列名
ORG = {"bu": None, "l2": None, "dept": None, "gender": None, "level": None, "perf": None, "tenure": None, "email": None}

def detect_schema(rows):
    """从首行表头识别题项列(Q1..Q12 按关键词)与组织列(一级/二级/三级)"""
    header = []
    for r in rows:
        header = [str(k).strip() for k in r.keys() if k not in (None, "")]
        break
    qdet = {}
    # 1) 精确匹配 Q1..Q12（支持标准格式短列名，大小写不敏感）
    for h in header:
        ht = h.strip()
        if len(ht) >= 2 and ht[0] in ("Q", "q") and ht[1:].isdigit():
            q = "Q" + ht[1:]
            if q in Q_KEYWORDS and q not in qdet:
                qdet[q] = h
    # 2) 题面关键词匹配（支持「我知道对我的工作要求。」这类完整题面表头）
    if len(qdet) < len(Q_KEYWORDS):
        for q, kws in Q_KEYWORDS.items():
            if q in qdet:
                continue
            for h in header:
                if any(kw in h for kw in kws):
                    qdet[q] = h
                    break
    org = {"bu": None, "l2": None, "dept": None, "gender": None, "level": None, "perf": None, "tenure": None, "email": None}
    for h in header:
        if org["dept"] is None and "三级" in h:
            org["dept"] = h
        elif org["l2"] is None and "二级" in h:
            org["l2"] = h
        elif org["bu"] is None and "一级" in h:
            org["bu"] = h
        elif org["gender"] is None and "性别" in h:
            org["gender"] = h
        elif org["level"] is None and "职级" in h:
            org["level"] = h
        elif org["perf"] is None and "绩效" in h:
            org["perf"] = h
        elif org["tenure"] is None and "司龄" in h:
            org["tenure"] = h
        elif org["email"] is None and ("邮箱" in h or "email" in h.lower()):
            org["email"] = h
    return qdet, org

# 5 分制阈值
BAND_RULES = [(4.5, "优势"), (4.0, "良好"), (3.5, "关注"), (-1, "预警")]
ENGAGED_MIN = 4.0     # 个人均分 >= 4.0 → 敬业
DISENGAGED_MAX = 3.0  # 个人均分 < 3.0 → 怠工
DEFAULT_L2 = "（直属团队）"  # 二级列缺省时的归并桶

def band(mean):
    if mean is None:
        return None
    for threshold, name in BAND_RULES:
        if mean >= threshold:
            return name
    return "预警"

def r2(x):
    return None if x is None else round(x, 2)

def load_rows(path):
    """读取 CSV/XLSX，返回 list[dict]"""
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        return [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]
    else:
        import csv
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                with open(p, newline="", encoding=enc) as f:
                    return [dict(r) for r in csv.DictReader(f)]
            except UnicodeDecodeError:
                continue
        raise SystemExit("无法识别 CSV 编码，请另存为 UTF-8")

def clean(rows):
    """校验并清洗，返回 (有效行, 警告列表)。固定 5 分制：分数原样保留"""
    valid, warnings = [], []
    for i, row in enumerate(rows, start=2):
        bu = str(row.get(ORG["bu"]) or "").strip()
        dept = str(row.get(ORG["dept"]) or "").strip()
        if not bu or not dept:
            warnings.append(f"第{i}行: 缺少 一级/三级 组织列，已跳过")
            continue
        l2 = str(row.get(ORG["l2"]) or "").strip() or DEFAULT_L2
        scores = {}
        for q in QPRESENT:
            v = row.get(QDETECT[q])
            if v is None or str(v).strip() == "":
                scores[q] = None  # 允许个别题缺答
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                warnings.append(f"第{i}行: {q} 非数字，按缺答处理")
                scores[q] = None
                continue
            if fv < 1 or fv > 5:
                warnings.append(f"第{i}行: {q}={fv} 超出1-5范围，按缺答处理")
                scores[q] = None
                continue
            scores[q] = fv
        if all(scores[q] is None for q in QPRESENT):
            warnings.append(f"第{i}行: 全部题目未作答，已跳过")
            continue
        valid.append({"bu": bu, "l2": l2, "dept": dept, "scores": scores,
                      "gender": str(row.get(ORG["gender"]) or "").strip() or None,
                      "level": str(row.get(ORG["level"]) or "").strip() or None,
                      "perf": str(row.get(ORG["perf"]) or "").strip() or None,
                      "tenure": str(row.get(ORG["tenure"]) or "").strip() or None,
                      "email": (str(row.get(ORG["email"])).strip() if ORG["email"] else None)})
    return valid, warnings

def agg_unit(members, min_n):
    """对一组员工计算完整指标; n < min_n 时 suppressed=True (分数仍计算但报告中须隐藏)"""
    n = len(members)
    unit = {"n": n, "suppressed": n < min_n}
    # 逐题
    q_stats = {}
    for q in QPRESENT:
        vals = [m["scores"][q] for m in members if m["scores"][q] is not None]
        q_stats[q] = {
            "mean": r2(sum(vals) / len(vals)) if vals else None,
            "n": len(vals),
            "low_pct": r2(100 * sum(1 for v in vals if v <= 2) / len(vals)) if vals else None,  # 打1-2分占比（不满）
            "high_pct": r2(100 * sum(1 for v in vals if v >= 4) / len(vals)) if vals else None,  # 打4-5分占比（认可）
            # 分数分布（5分制三段）：用于经理人看板判断"共识偏低"还是"两极分化"
            "dist": ({"low": sum(1 for v in vals if v <= 2),  # 1-2 分：不满
                      "mid": sum(1 for v in vals if v == 3),   # 3 分：中性
                      "high": sum(1 for v in vals if v >= 4)} # 4-5 分：认可
                     if vals else None),
        }
        q_stats[q]["band"] = band(q_stats[q]["mean"])
    unit["questions"] = q_stats
    # 维度
    dims = {}
    for dname, dcfg in DIMENSIONS.items():
        means = [q_stats[q]["mean"] for q in dcfg["questions"]
                if q in QPRESENT and q_stats[q]["mean"] is not None]
        dm = sum(means) / len(means) if means else None
        dims[dname] = {"mean": r2(dm), "band": band(dm)}
    unit["dimensions"] = dims
    # 总均值 (实际题项均值的均值)
    all_means = [q_stats[q]["mean"] for q in QPRESENT if q_stats[q]["mean"] is not None]
    gm = sum(all_means) / len(all_means) if all_means else None
    unit["grand_mean"] = r2(gm)
    unit["grand_band"] = band(gm)
    # 敬业度分布（按个人均分）
    engaged = neutral = disengaged = 0
    for m in members:
        vals = [v for v in m["scores"].values() if v is not None]
        pm = sum(vals) / len(vals)
        if pm >= ENGAGED_MIN:
            engaged += 1
        elif pm < DISENGAGED_MAX:
            disengaged += 1
        else:
            neutral += 1
    unit["engagement"] = {
        "engaged_pct": r2(100 * engaged / n),
        "neutral_pct": r2(100 * neutral / n),
        "disengaged_pct": r2(100 * disengaged / n),
        "ratio": r2(engaged / disengaged) if disengaged else None,  # 敬业:怠工比
    }
    # 人员间标准差（个人均分的标准差），用于 VP 看板「经理人效能象限」
    pms = [sum(v for v in m["scores"].values() if v is not None) /
           sum(1 for v in m["scores"].values() if v is not None)
           for m in members]
    unit["score_std"] = r2(statistics.stdev(pms)) if len(pms) > 1 else 0.0
    # 最高/最低题项
    ranked = sorted([q for q in QPRESENT if q_stats[q]["mean"] is not None],
                    key=lambda q: q_stats[q]["mean"])
    unit["bottom_questions"] = ranked[:2]
    unit["top_questions"] = ranked[-2:][::-1]
    # 人口学分组聚合 + 管理者流失风险名单
    unit["demographics"] = agg_demographics(members)
    unit["managers"] = agg_managers(members)
    return unit


def _pm(m):
    """个人均分（该员工所有有效题项均值）"""
    vals = [v for v in m["scores"].values() if v is not None]
    return sum(vals) / len(vals) if vals else None


def _level_num(level):
    if not level:
        return 0
    m = re.search(r"\d+", str(level))
    return int(m.group()) if m else 0


def agg_demographics(members):
    """按 性别 / 职级 / 司龄 / 绩效 分组的健康度聚合（用于 VP 看板「员工群体洞察」）"""
    fields = {"gender": "gender", "level": "level", "tenure": "tenure", "perf": "perf"}
    out = {}
    for gname, attr in fields.items():
        buckets = {}
        for m in members:
            k = m.get(attr)
            if not k:
                continue
            buckets.setdefault(k, []).append(m)
        grp = {}
        for k, ms in buckets.items():
            pms = [v for v in (_pm(m) for m in ms) if v is not None]
            n = len(ms)
            if not pms:
                continue
            gm = sum(pms) / len(pms)
            eng = sum(1 for v in pms if v >= ENGAGED_MIN)
            dis = sum(1 for v in pms if v < DISENGAGED_MAX)
            grp[k] = {
                "n": n,
                "grand_mean": r2(gm),
                "health_index": r2(gm / 5 * 100),
                "engaged_pct": r2(100 * eng / n),
                "disengaged_pct": r2(100 * dis / n),
            }
        out[gname] = grp
    return out


def agg_managers(members):
    """P7+ 管理者流失风险名单（用于 VP 看板「管理者流失风险分析」）。
    健康度 = 个人均分/5*100；<60 高危 / 60-80 需关注 / >=80 稳定。"""
    mg = [m for m in members if _level_num(m.get("level")) >= 7]
    out = []
    for m in mg:
        pmv = _pm(m)
        hi = round(pmv / 5 * 100, 1) if pmv is not None else None
        if hi is None or hi >= 80:
            risk = "稳定"
        elif hi >= 60:
            risk = "需关注"
        else:
            risk = "高危"
        out.append({
            "email": m.get("email"), "level": m.get("level"), "tenure": m.get("tenure"),
            "perf": m.get("perf"), "mean": r2(pmv), "health_index": hi, "risk": risk,
        })
    out.sort(key=lambda x: (x["health_index"] is None, x["health_index"] or 0))  # 健康度升序 → 风险高在前
    total = len(out)
    summary = {
        "total": total,
        "high_risk": sum(1 for x in out if x["risk"] == "高危"),
        "watch": sum(1 for x in out if x["risk"] == "需关注"),
        "stable": sum(1 for x in out if x["risk"] == "稳定"),
        "risk_rate": r2(100 * sum(1 for x in out if x["risk"] == "高危") / total) if total else None,
    }
    return {"summary": summary, "managers": out}

def pct_rank(value, population):
    """value 在 population 中的百分位 (0-100, 打败了多少同类单位)"""
    others = [p for p in population if p is not None]
    if value is None or len(others) <= 1:
        return None
    below = sum(1 for p in others if p < value)
    return round(100 * below / len(others))

def attach_history(unit, prev_unit):
    """把上期均值与变化写入 unit"""
    if not prev_unit:
        return
    pg = prev_unit.get("grand_mean")
    if pg is not None and unit.get("grand_mean") is not None:
        unit["grand_mean_prev"] = pg
        unit["grand_mean_delta"] = r2(unit["grand_mean"] - pg)
    for q in QPRESENT:
        pm = (prev_unit.get("questions") or {}).get(q, {}).get("mean")
        if pm is not None and unit["questions"][q]["mean"] is not None:
            unit["questions"][q]["prev"] = pm
            unit["questions"][q]["delta"] = r2(unit["questions"][q]["mean"] - pm)
    for d in DIMENSIONS:
        pm = (prev_unit.get("dimensions") or {}).get(d, {}).get("mean")
        if pm is not None and unit["dimensions"][d]["mean"] is not None:
            unit["dimensions"][d]["prev"] = pm
            unit["dimensions"][d]["delta"] = r2(unit["dimensions"][d]["mean"] - pm)

def prev_dept(prev, bu, l2, dept):
    return (((prev.get("business_units", {}).get(bu, {})
              .get("l2_units", {}).get(l2, {})
              .get("departments", {}) or {}).get(dept)) if prev else None)

def prev_l2(prev, bu, l2):
    return (((prev.get("business_units", {}).get(bu, {})
              .get("l2_units", {}) or {}).get(l2)) if prev else None)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="员工明细表 CSV/XLSX")
    ap.add_argument("--output", default="analysis.json")
    ap.add_argument("--history", help="上期 analysis.json (可选)")
    ap.add_argument("--period", default="", help='本期标签，如 "2026 H1"')
    ap.add_argument("--min-n", type=int, default=4, help="数据抑制阈值，低于此人数的单位不显示分数")
    args = ap.parse_args()

    rows = load_rows(args.input)
    qdet, org = detect_schema(rows)
    if not qdet:
        raise SystemExit("未识别到任何问卷题项，请检查表头是否为标准题面文字")
    if not (org["bu"] and org["dept"]):
        raise SystemExit("未识别到 一级/三级 组织列，请检查表头（应包含「一级」「三级」字样）")
    global QPRESENT, QDETECT, ORG
    QPRESENT = sorted(qdet.keys(), key=lambda q: int(q[1:]))
    QDETECT = qdet
    ORG = org
    members, warnings = clean(rows)
    if not members:
        raise SystemExit("没有有效数据行")

    prev = None
    if args.history:
        prev = json.loads(Path(args.history).read_text(encoding="utf-8"))

    # 公司整体
    company = agg_unit(members, args.min_n)
    attach_history(company, prev.get("company") if prev else None)

    # 一级事业部 → 二级 → 三级部门
    bus = {}
    bu_names = sorted({m["bu"] for m in members})
    for bu in bu_names:
        bu_members = [m for m in members if m["bu"] == bu]
        bu_unit = agg_unit(bu_members, args.min_n)
        l2_units = {}
        for l2 in sorted({m["l2"] for m in bu_members}):
            l2_members = [m for m in bu_members if m["l2"] == l2]
            l2_unit = agg_unit(l2_members, args.min_n)
            depts = {}
            for dept in sorted({m["dept"] for m in l2_members}):
                d_unit = agg_unit([m for m in l2_members if m["dept"] == dept], args.min_n)
                attach_history(d_unit, prev_dept(prev, bu, l2, dept))
                depts[dept] = d_unit
            l2_unit["departments"] = depts
            attach_history(l2_unit, prev_l2(prev, bu, l2))
            l2_units[l2] = l2_unit
        bu_unit["l2_units"] = l2_units
        attach_history(bu_unit, (prev or {}).get("business_units", {}).get(bu) if prev else None)
        bus[bu] = bu_unit

    # 内部百分位：同级单位排名（公司全量对比）
    bu_gms = [bus[b]["grand_mean"] for b in bus]
    for b in bus:
        bus[b]["percentile_vs_bus"] = pct_rank(bus[b]["grand_mean"], bu_gms)

    all_l2_gms = [l2["grand_mean"] for b in bus for l2 in bus[b]["l2_units"].values()]
    for b in bus:
        for l2 in bus[b]["l2_units"].values():
            l2["percentile_vs_l2"] = (None if l2["suppressed"]
                                       else pct_rank(l2["grand_mean"], all_l2_gms))

    all_dept_gms = [d["grand_mean"]
                    for b in bus for l2 in bus[b]["l2_units"].values()
                    for d in l2["departments"].values() if not d["suppressed"]]
    for b in bus:
        for l2 in bus[b]["l2_units"].values():
            for d in l2["departments"].values():
                d["percentile_vs_depts"] = (None if d["suppressed"]
                                            else pct_rank(d["grand_mean"], all_dept_gms))

    # —— 系统性 vs 局部问题诊断（CEO视角）：判定每题是全司共性(政策级)还是局部(管理辅导)
    qdiag = {}
    for q in QPRESENT:
        cm = company["questions"][q]["mean"]
        bu_means = [bus[b]["questions"][q]["mean"] for b in bus
                    if not bus[b]["suppressed"] and bus[b]["questions"][q]["mean"] is not None]
        total = len(bu_means)
        below = sum(1 for m in bu_means if m < 3.5)
        if cm is not None and cm >= 4.5:
            qtype = "strength"
        elif cm is not None and cm < 4.0 and total > 0 and below / total >= 0.5:
            qtype = "systemic"
        else:
            qtype = "localized"
        qdiag[q] = {"type": qtype, "company_mean": r2(cm),
                    "bu_below": below, "bu_total": total}
    company["question_diagnosis"] = qdiag

    # —— 主管效能标签（VP视角）：在各事业部内对三级部门负责人排名并标注
    for b in bus:
        bu = bus[b]
        dept_list = [(dn, d) for l2 in bu["l2_units"].values()
                     for dn, d in l2["departments"].items() if not d["suppressed"]]
        dept_list.sort(key=lambda kv: (kv[1]["grand_mean"] is None, -(kv[1]["grand_mean"] or 0)))
        n_dept = len(dept_list)
        for i, (dn, d) in enumerate(dept_list):
            gm = d["grand_mean"]
            dis = d["engagement"]["disengaged_pct"]
            if gm is not None and gm >= 4.5:
                tag = "优秀标杆"
            elif gm is not None and gm < 3.5:
                tag = "需辅导"
            elif dis >= 20:
                tag = "需辅导"
            elif n_dept >= 4 and i < max(1, n_dept // 4):
                tag = "优秀标杆"
            else:
                tag = "稳健"
            d["manager_tag"] = tag
            d["rank_in_bu"] = i + 1
            d["n_depts_in_bu"] = n_dept

    total_dept = sum(len(l2["departments"]) for b in bus for l2 in bus[b]["l2_units"].values())
    total_l2 = sum(len(bus[b]["l2_units"]) for b in bus)
    result = {
        "meta": {
            "period": args.period,
            "total_respondents": len(members),
            "min_n": args.min_n,
            "has_history": prev is not None,
            "prev_period": (prev or {}).get("meta", {}).get("period", "") if prev else "",
            "scale": "1-5",
            "n_questions": len(QPRESENT),
            "questions": {q: QUESTIONS[q] for q in QPRESENT},
            "question_short": {q: Q_SHORT[q] for q in QPRESENT},
            "dimensions": {k: {"questions": [q for q in v["questions"] if q in QPRESENT],
                            "core_question": v["core_question"]} for k, v in DIMENSIONS.items()},
            "band_rules": "均值>=4.5 优势 | 4.0-4.49 良好 | 3.5-3.99 关注 | <3.5 预警",
            "warnings": warnings,
        },
        "company": company,
        "business_units": bus,
    }
    Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"OK 已生成 {args.output}")
    print(f"  受访人数: {len(members)} | 一级事业部: {len(bus)} | "
          f"二级单位: {total_l2} | 三级部门: {total_dept}")
    print(f"  公司总均值: {company['grand_mean']} ({company['grand_band']})")
    if warnings:
        print(f"  警告 {len(warnings)} 条(已写入 meta.warnings)，前5条:")
        for w in warnings[:5]:
            print("   -", w)

if __name__ == "__main__":
    main()
